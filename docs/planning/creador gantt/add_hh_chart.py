#!/usr/bin/env python3
"""
HERMES-A1 — Add monthly HH bar chart to planning xlsx
Creates a second sheet "HH Mensuales" with a grouped bar chart:
- X axis: weeks grouped by month
- Y axis: hours (HH)
- Max 15 HH/week line shown
"""
import openpyxl, re, os, glob, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

PLANNING_DIR = os.path.dirname(os.path.abspath(__file__))
matches = glob.glob(os.path.join(PLANNING_DIR, 'Planif*'))
if not matches:
    raise RuntimeError("No planning file found")
DST_USER = matches[0]

# ═══════════════════════════════════════════════════════════════════════
# STEP 1: Read existing data
# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.load_workbook(DST_USER)
ws = wb.active

months_es = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'Abril':4,'May':5,'Jun':6,
             'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}

def parse_week_date(title):
    months_re = '|'.join(months_es.keys())
    m = re.search(r'(\d{1,2})\s+(' + months_re + r')\s+al\s', title)
    if m:
        day, month_s = int(m.group(1)), m.group(2)
        ym = re.search(r'(\d{4})', title)
        year = int(ym.group(1)) if ym else 2026
        return datetime(year, months_es.get(month_s, 1), day)
    m = re.search(r'(\d{1,2})\s+al\s+(\d{1,2})\s+(\w+)', title)
    if m:
        day, month_s = int(m.group(1)), m.group(3)
        ym = re.search(r'(\d{4})', title)
        year = int(ym.group(1)) if ym else 2026
        return datetime(year, months_es.get(month_s, 1), day)
    return datetime(2025, 1, 1)

def parse_hh(text):
    m = re.search(r'\[(\d+)\s*HH\]', str(text or ''))
    return int(m.group(1)) if m else 0

# Extract week data
all_weeks = []
current_week_title = None
current_week_items = []

for r in range(1, ws.max_row + 1):
    c1_val = ws.cell(r, 1).value
    c5_val = ws.cell(r, 5).value
    c1_str = str(c1_val) if c1_val else ''
    
    if 'Semana' in c1_str and ':' in c1_str:
        if current_week_title:
            all_weeks.append((current_week_title, current_week_items[:]))
        current_week_title = c1_str.strip()
        current_week_items = []
    elif c1_str in ('1', '2', '3', '4', '5') and current_week_title:
        hh = parse_hh(c5_val)
        current_week_items.append(hh)
    elif 'Total HH' in c1_str and current_week_title:
        tm = re.search(r'Total HH:\s*(\d+)', c1_str)
        total = int(tm.group(1)) if tm else sum(current_week_items)
        all_weeks.append((current_week_title, current_week_items[:], total))
        current_week_title = None
        current_week_items = []

if current_week_title:
    all_weeks.append((current_week_title, current_week_items[:], sum(current_week_items)))

print('Found %d weeks' % len(all_weeks))

# Sort chronologically (oldest first for chart)
week_data = []
for entry in all_weeks:
    title = entry[0]
    items = entry[1]
    total = entry[2] if len(entry) > 2 else sum(items)
    dt = parse_week_date(title)
    # Clean title for chart label
    clean = re.sub(r'\[.*?\]', '', title).strip()
    # Short label: "Sem 26 (23-27 Jun)"
    sem_m = re.search(r'Semana\s+(\S+)', clean)
    date_m = re.search(r'(\d{1,2})\s+al\s+(\d{1,2})\s+(\w+)', clean)
    if sem_m and date_m:
        short = 'Sem %s\n(%s-%s %s)' % (sem_m.group(1), date_m.group(1), date_m.group(2), date_m.group(3))
    else:
        short = clean[:25]
    is_actual = '[ACTUAL]' in title or '[SESION' in title.upper()
    week_data.append({
        'title': title,
        'short': short,
        'date': dt,
        'items': items,
        'total': total,
        'is_actual': is_actual,
    })

# Sort oldest first
week_data.sort(key=lambda w: w['date'])

# Group by month
by_month = defaultdict(list)
for wd in week_data:
    key = (wd['date'].year, wd['date'].month)
    by_month[key].append(wd)

months_order = sorted(by_month.keys())

print('\nMonths found:')
for ym in months_order:
    month_name = datetime(ym[0], ym[1], 1).strftime('%B %Y')
    weeks_in_month = by_month[ym]
    total_hh = sum(w['total'] for w in weeks_in_month)
    print('  %s: %d weeks, %d HH total' % (month_name, len(weeks_in_month), total_hh))

# ═══════════════════════════════════════════════════════════════════════
# STEP 2: Create chart sheet
# ═══════════════════════════════════════════════════════════════════════
if 'HH Mensuales' in wb.sheetnames:
    del wb['HH Mensuales']
ws2 = wb.create_sheet('HH Mensuales')

# Colors
DK_BLUE = "1F3864"
MD_BLUE = "2E75B6"
LT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LT_GRAY = "F2F2F2"
DK_GRAY = "404040"
GREEN = "70AD47"
LT_GREEN = "E2EFDA"
ORANGE = "ED7D31"
AMBER = "FFC000"
RED = "C00000"

f_title = PatternFill("solid", fgColor=DK_BLUE)
f_subtitle = PatternFill("solid", fgColor="2B4C7E")
f_colhdr = PatternFill("solid", fgColor=MD_BLUE)
f_cell = PatternFill("solid", fgColor=LT_BLUE)
f_cell2 = PatternFill("solid", fgColor=LT_GRAY)
f_over = PatternFill("solid", fgColor="FCE4D6")
f_actual = PatternFill("solid", fgColor=LT_GREEN)
f_month_hdr = PatternFill("solid", fgColor=DK_BLUE)

fn_title = Font(name="Calibri", bold=True, size=14, color=WHITE)
fn_subtitle = Font(name="Calibri", size=10, color=WHITE)
fn_hdr = Font(name="Calibri", bold=True, size=10, color=WHITE)
fn_cell = Font(name="Calibri", size=10, color=DK_GRAY)
fn_cell_bold = Font(name="Calibri", bold=True, size=10, color=DK_GRAY)
fn_over = Font(name="Calibri", bold=True, size=10, color=RED)
fn_ok = Font(name="Calibri", bold=True, size=10, color=GREEN)
fn_actual = Font(name="Calibri", bold=True, size=10, color="2B6B1E")
fn_footer = Font(name="Calibri", italic=True, size=9, color=DK_GRAY)

thin = Side(style="thin", color="B4B4B4")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
al = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Title rows ──────────────────────────────────────────────────
ws2.column_dimensions['A'].width = 4
ws2.column_dimensions['B'].width = 28

ws2.merge_cells('A1:E1')
c = ws2.cell(1, 1, value="HERMES-A1 — Horas Hombre Mensuales por Semana")
c.font = fn_title; c.fill = f_title; c.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 32
for cc in range(1, 6):
    ws2.cell(1, cc).fill = f_title; ws2.cell(1, cc).border = bd

ws2.merge_cells('A2:E2')
c = ws2.cell(2, 1, value="Grafico de barras agrupado por mes | Maximo: 15 HH/semana | LIND SpA | Junio 2026")
c.font = fn_subtitle; c.fill = f_subtitle; c.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[2].height = 20
for cc in range(1, 6):
    ws2.cell(2, cc).fill = f_subtitle; ws2.cell(2, cc).border = bd

# ── Data Table ──────────────────────────────────────────────────
# Layout: Row 4 = month headers, Row 5 = week labels, Rows 6+ = data
data_start_row = 4

r = data_start_row

# For each month group, write headers
chart_col = 2  # Start at B

# First: write a flat table with all weeks as columns
# Row 4: Month group headers (merged)
# Row 5: Week short labels
# Row 6: HH values
# Row 7: Status (OK/OVER)

# Build flat list of all weeks (oldest first)
all_sorted = sorted(week_data, key=lambda w: w['date'])
num_weeks = len(all_sorted)

# Write month group headers (row 4)
month_col_start = 2
for ym in months_order:
    month_name = datetime(ym[0], ym[1], 1).strftime('%b %Y')
    weeks_in = by_month[ym]
    n_weeks = len(weeks_in)
    month_col_end = month_col_start + n_weeks - 1
    
    if n_weeks > 1:
        ws2.merge_cells(start_row=r, start_column=month_col_start, end_row=r, end_column=month_col_end)
    c = ws2.cell(r, month_col_start, value=month_name)
    c.font = fn_hdr; c.fill = f_month_hdr; c.alignment = ac
    for cc in range(month_col_start, month_col_end + 1):
        ws2.cell(r, cc).fill = f_month_hdr; ws2.cell(r, cc).border = bd
        ws2.column_dimensions[get_column_letter(cc)].width = 16
    month_col_start = month_col_end + 1

ws2.row_dimensions[r].height = 22
r += 1  # Row 5: week labels

# Week short labels
col = 2
week_col_map = {}  # map week index -> column
for idx, wd in enumerate(all_sorted):
    c = ws2.cell(r, col, value=wd['short'])
    if wd['is_actual']:
        c.font = fn_actual; c.fill = f_actual
    else:
        c.font = fn_cell_bold; c.fill = f_cell
    c.alignment = ac; c.border = bd
    week_col_map[idx] = col
    col += 1

ws2.row_dimensions[r].height = 42
r += 1  # Row 6: HH values

# HH values
hh_row = r
col = 2
for idx, wd in enumerate(all_sorted):
    total = wd['total']
    c = ws2.cell(r, col, value=total)
    if total > 15:
        c.font = fn_over; c.fill = f_over
    elif total > 0:
        c.font = fn_ok; c.fill = f_cell2
    else:
        c.font = fn_cell; c.fill = f_cell2
    c.alignment = ac; c.border = bd
    col += 1

ws2.row_dimensions[r].height = 22
r += 1  # Row 7: status

# Status row
col = 2
for idx, wd in enumerate(all_sorted):
    total = wd['total']
    if total > 15:
        status = 'OVER +%d' % (total - 15)
    elif total == 0:
        status = 'Sin HH'
    else:
        status = 'OK'
    c = ws2.cell(r, col, value=status)
    if total > 15:
        c.font = fn_over; c.fill = f_over
    elif total == 0:
        c.font = fn_cell; c.fill = f_cell2
    else:
        c.font = fn_ok; c.fill = f_cell2
    c.alignment = ac; c.border = bd
    col += 1

ws2.row_dimensions[r].height = 18
r += 1  # Row 8: blank
ws2.row_dimensions[r].height = 6

# ── Monthly totals row ──────────────────────────────────────────
r_total = r
ws2.cell(r, 1, value="").border = bd
col = 2
for ym in months_order:
    weeks_in = by_month[ym]
    n_weeks = len(weeks_in)
    total_month = sum(w['total'] for w in weeks_in)
    avg_month = total_month / n_weeks if n_weeks > 0 else 0
    
    if n_weeks > 1:
        ws2.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + n_weeks - 1)
    
    c = ws2.cell(r, col, value='%d HH total' % total_month)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=MD_BLUE)
    c.alignment = ac
    for cc in range(col, col + n_weeks):
        ws2.cell(r, cc).fill = PatternFill("solid", fgColor=MD_BLUE)
        ws2.cell(r, cc).border = bd
    col += n_weeks

ws2.row_dimensions[r].height = 22
r += 2  # skip a line

# ── Create Bar Chart ──────────────────────────────────────────
chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = "Horas Hombre por Semana"
chart.y_axis.title = "Horas Hombre (HH)"
chart.x_axis.title = "Semana"
chart.style = 10
chart.width = 30
chart.height = 16

# Data = HH values row
data_ref = Reference(ws2, min_col=2, min_row=hh_row, max_col=1 + num_weeks, max_row=hh_row)

# Categories = week labels
cats_ref = Reference(ws2, min_col=2, min_row=hh_row - 1, max_col=1 + num_weeks, max_row=hh_row - 1)

chart.add_data(data_ref, from_rows=True, titles_from_data=False)
chart.set_categories(cats_ref)

# Series name
chart.series[0].title = openpyxl.chart.series.SeriesLabel(v="HH por Semana")

# Colors
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

# Color each bar based on actual/over/normal
for idx, wd in enumerate(all_sorted):
    pt = DataPoint(idx=idx)
    if wd['is_actual']:
        pt.graphicalProperties.solidFill = GREEN
    elif wd['total'] > 15:
        pt.graphicalProperties.solidFill = ORANGE
    elif wd['total'] == 0:
        pt.graphicalProperties.solidFill = LT_GRAY
    else:
        pt.graphicalProperties.solidFill = MD_BLUE
    chart.series[0].data_points.append(pt)

# Data labels
chart.series[0].dLbls = DataLabelList()
chart.series[0].dLbls.showVal = True
chart.series[0].dLbls.numFmt = '0'

# Max line: add a second series with value 15 for each week
max_data = [15] * num_weeks
from openpyxl.chart import LineChart
line_chart = LineChart()
line_data = Reference(ws2, min_col=2, min_row=r_total - 1, max_col=1 + num_weeks, max_row=r_total - 1)
# Actually, let's just create a separate row with 15s for the max line
max_row = r_total - 1
ws2.cell(max_row, 1, value="Max").font = fn_cell
for col_idx in range(2, 2 + num_weeks):
    ws2.cell(max_row, col_idx, value=15)

max_ref = Reference(ws2, min_col=2, min_row=max_row, max_col=1 + num_weeks, max_row=max_row)
line_chart.add_data(max_ref, from_rows=True, titles_from_data=False)
line_chart.series[0].title = openpyxl.chart.series.SeriesLabel(v="Max 15 HH")
line_chart.series[0].graphicalProperties.solidFill = RED
line_chart.series[0].graphicalProperties.line.solidFill = RED
line_chart.series[0].graphicalProperties.line.width = 25000  # 2pt

# Combine
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 20
chart += line_chart

# Place chart below data
chart_row = r + 1
ws2.add_chart(chart, "A%d" % chart_row)

# ── Footer ──────────────────────────────────────────────────
footer_row = chart_row + 20
ws2.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)
c = ws2.cell(footer_row, 1, value="Generado automaticamente | Actualizado: 24 Junio 2026 | LIND SpA")
c.font = fn_footer; c.alignment = Alignment(horizontal="center", vertical="center")

# ── Summary table below chart ──────────────────────────────────
summary_row = footer_row + 2
ws2.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
c = ws2.cell(summary_row, 1, value="Resumen por Mes")
c.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
c.fill = f_title; c.alignment = Alignment(horizontal="left", vertical="center")
for cc in range(1, 4):
    ws2.cell(summary_row, cc).fill = f_title; ws2.cell(summary_row, cc).border = bd
ws2.row_dimensions[summary_row].height = 24

summary_row += 1
# Headers
for ci, hdr in enumerate(['Mes', 'Semanas', 'HH Total', 'HH Promedio'], 1):
    c = ws2.cell(summary_row, ci, value=hdr)
    c.font = fn_hdr; c.fill = f_colhdr; c.alignment = ac; c.border = bd
ws2.row_dimensions[summary_row].height = 22

summary_row += 1
total_all_hh = 0
total_all_weeks = 0
for ym in months_order:
    month_name = datetime(ym[0], ym[1], 1).strftime('%B %Y')
    weeks_in = by_month[ym]
    n_weeks = len(weeks_in)
    total_hh = sum(w['total'] for w in weeks_in)
    avg_hh = total_hh / n_weeks if n_weeks > 0 else 0
    total_all_hh += total_hh
    total_all_weeks += n_weeks
    
    ws2.cell(summary_row, 1, value=month_name).font = fn_cell
    ws2.cell(summary_row, 1).fill = f_cell; ws2.cell(summary_row, 1).alignment = al; ws2.cell(summary_row, 1).border = bd
    
    ws2.cell(summary_row, 2, value=n_weeks).font = fn_cell
    ws2.cell(summary_row, 2).fill = f_cell2; ws2.cell(summary_row, 2).alignment = ac; ws2.cell(summary_row, 2).border = bd
    
    ws2.cell(summary_row, 3, value=total_hh).font = fn_cell_bold
    ws2.cell(summary_row, 3).fill = f_cell2; ws2.cell(summary_row, 3).alignment = ac; ws2.cell(summary_row, 3).border = bd
    
    ws2.cell(summary_row, 4, value=round(avg_hh, 1)).font = fn_cell
    ws2.cell(summary_row, 4).fill = f_cell; ws2.cell(summary_row, 4).alignment = ac; ws2.cell(summary_row, 4).border = bd
    
    ws2.row_dimensions[summary_row].height = 22
    summary_row += 1

# Total row
for ci, val in enumerate(['TOTAL', total_all_weeks, total_all_hh, round(total_all_hh / total_all_weeks, 1) if total_all_weeks > 0 else 0], 1):
    c = ws2.cell(summary_row, ci, value=val)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=MD_BLUE)
    c.alignment = ac; c.border = bd
ws2.row_dimensions[summary_row].height = 22

# ═══════════════════════════════════════════════════════════════════════
# STEP 3: Save
# ═══════════════════════════════════════════════════════════════════════
import shutil

tmp = DST_USER + '.tmp.xlsx'
wb.save(tmp)
print('Saved temp: %d bytes' % os.path.getsize(tmp))

try: os.remove(DST_USER)
except (PermissionError, FileNotFoundError): pass

shutil.move(tmp, DST_USER)
print('Final: %s (%d bytes)' % (os.path.basename(DST_USER), os.path.getsize(DST_USER)))
print('Sheet "HH Mensuales" added with %d weeks bar chart + summary table' % num_weeks)
