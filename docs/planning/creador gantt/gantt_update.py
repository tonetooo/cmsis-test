#!/usr/bin/env python3
"""
HERMES A1 — Carta Gantt 2026
Generates a fully updated Gantt chart from the original 2025 template.
Preserves header styling, replaces all data with current project status
and future roadmap.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from copy import copy
import os

SRC = "../creador gantt/GANTT 2025.xlsx"
DST = os.path.join(os.path.dirname(os.path.abspath(__file__)), "GANTT 2026.xlsx")

# ── Colours ──────────────────────────────────────────────────────────────
GREEN      = "70AD47"
DARK_GREEN = "00B050"
BLUE       = "5B9BD5"
DARK_BLUE  = "2F5496"
ORANGE     = "ED7D31"
AMBER      = "FFC000"
YELLOW     = "FFF2CC"
RED        = "FF0000"
WHITE      = "FFFFFF"
LIGHT_GRAY = "D9D9D9"
DARK_GRAY  = "404040"
PHASE_BG   = "1F4E79"
PHASE_FG   = "FFFFFF"
MILESTONE  = "BDD7EE"

fills = {
    "green":     PatternFill("solid", fgColor=GREEN),
    "darkgreen": PatternFill("solid", fgColor=DARK_GREEN),
    "blue":      PatternFill("solid", fgColor=BLUE),
    "darkblue":  PatternFill("solid", fgColor=DARK_BLUE),
    "orange":    PatternFill("solid", fgColor=ORANGE),
    "amber":     PatternFill("solid", fgColor=AMBER),
    "yellow":    PatternFill("solid", fgColor=YELLOW),
    "lightgray": PatternFill("solid", fgColor=LIGHT_GRAY),
    "white":     PatternFill("solid", fgColor=WHITE),
    "phase":     PatternFill("solid", fgColor=PHASE_BG),
    "milestone": PatternFill("solid", fgColor=MILESTONE),
    "red":       PatternFill("solid", fgColor=RED),
}

thin_border = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

header_font = Font(name="Calibri", bold=True, size=10, color=WHITE)
phase_font  = Font(name="Calibri", bold=True, size=10, color=WHITE)
task_font   = Font(name="Calibri", size=9)
pct_font    = Font(name="Calibri", size=9, bold=True)
date_font   = Font(name="Calibri", size=8, color=DARK_GRAY)
milestone_font = Font(name="Calibri", bold=True, size=9, color=DARK_BLUE)
title_font  = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
subtitle_font = Font(name="Calibri", size=10, color=DARK_GRAY)

# ── Timeline: monthly columns Jan 2025 → Dec 2026 ───────────────────────
month_labels = []
curr = date(2025, 1, 1)
end = date(2026, 12, 1)
MONTHS_ES = [
    "Ene","Feb","Mar","Abr","May","Jun",
    "Jul","Ago","Sep","Oct","Nov","Dic"
]
while curr <= end:
    month_labels.append(f"{MONTHS_ES[curr.month-1]} {curr.year}")
    # next month
    y = curr.year + (curr.month // 12)
    m = (curr.month % 12) + 1
    if m == 1:
        curr = date(y, 1, 1)
    else:
        curr = date(curr.year, m, 1)
NUM_MONTHS = len(month_labels)  # 24 months


# ── Phase / Task database ───────────────────────────────────────────────
# Each entry: (type, phase_label|task_name, assigned, progress%, start_m, end_m)
# start_m / end_m = 0-based month index (0 = Jan 2025, 23 = Dec 2026)
# Bar colour derived from progress: 100=green, >=75=blue, >=25=orange, >0=amber, 0=none
# Phase entries use fills["phase"]

class Task:
    __slots__ = ("kind","label","assigned","pct","sm","em","note")
    def __init__(self, kind, label="", assigned="", pct=0, sm=0, em=0, note=""):
        self.kind = kind       # "phase","task","milestone","sep"
        self.label = label
        self.assigned = assigned
        self.pct = pct
        self.sm = sm          # start month index
        self.em = em          # end month index
        self.note = note

ROWS = [
    # ═══════════════════════════════════════════════════════════════════
    Task("phase", "✅ LOGRADO — FASE 1: LOGÍSTICA Y HARDWARE (Ene 2025)"),
    Task("task", "Gestión de Compra Sixfab y Mouser", "Antonio A.", 100, 0, 0),
    Task("task", "Pruebas de Encendido y Voltajes con RPi5", "Antonio A.", 100, 0, 0),
    Task("task", "Recepción y Montaje Físico", "Antonio A.", 100, 0, 0, "Chasis + conexiones"),
    Task("task", "Integración HAT Sixfab (UART 3.3V nativo)", "Antonio A.", 100, 0, 1),

    # ═══════════════════════════════════════════════════════════════════
    Task("phase", "✅ LOGRADO — FASE 2: FIRMWARE Y PORTING (Ene-Jun 2025)"),
    Task("task", "Porting Algoritmo 'Dos Golpes'", "Antonio A.", 100, 0, 0),
    Task("task", "Desarrollo Filtro 'Level to Zero'", "Antonio A.", 100, 1, 1),
    Task("task", "Implementación Almacenamiento Local (SDIO/MicroSD)", "Antonio A.", 100, 0, 1),
    Task("task", "Máquina de Estados UART / Comandos AT", "Antonio A.", 100, 0, 0),
    Task("task", "Integración Stack LTE (Quectel EC25)", "Antonio A.", 100, 0, 0),

    # ═══════════════════════════════════════════════════════════════════
    Task("phase", "✅ LOGRADO — FASE 3: MIGRACIÓN A FreeRTOS (Abr-May 2026)"),
    Task("task", "Migración a FreeRTOS + CMSIS-RTOS v2", "Antonio A.", 100, 15, 16),
    Task("task", "5 Tareas RTOS (sensor, modem, file, control, modem_sim)",
         "Antonio A.", 100, 15, 16),
    Task("task", "Driver SPI2 con DMA para ADXL355", "Antonio A.", 100, 15, 16,
         "DMA1 Stream 3/4, circular + FIFO"),
    Task("task", "SD/FatFs + CLI (sdtest, mount, format)", "Antonio A.", 100, 15, 16),
    Task("task", "Módulos Algoritmo (sensor_algo, cli_algo, csv_algo)",
         "Antonio A.", 100, 16, 17),
    Task("task", "Watchdog IWDG implementado y probado", "Antonio A.", 100, 16, 16),
    Task("task", "Suite de Pruebas (Ceedling + Unity, 47 tests)", "Antonio A.", 100, 16, 17),

    # ═══════════════════════════════════════════════════════════════════
    Task("sep"),
    Task("phase", "📋 P1 — SISTEMA DE ADQUISICIÓN (Prioridad Alta · ~6 meses)"),
    Task("task", "Adquisición en Tiempo Continuo (sin cortes)", "Antonio A.", 85, 17, 22,
         "Pipeline con cola RAM + adquisición concurrente con upload"),
    Task("task", "Optimización DMA (completar validación + modo burst)", "Antonio A.", 60, 17, 19),
    Task("task", "Calibración y Caracterización Acelerómetro", "Antonio A.", 0, 17, 20),
    Task("task", "Buffer Local de Datos ante pérdida de conectividad",
         "Antonio A.", 90, 18, 21,
         "Cola RAM 8 slots + backup registers RTC BKP0R-BKP19R (survive NRST)"),
    Task("task", "Sincronización Horaria (NTP sobre LTE)", "Antonio A.", 10, 18, 21,
         "AT+QNTP implementado pero falla con ambos servidores"),
    Task("task", "FPU Enable (configENABLE_FPU=1) + migración matemática",
         "Antonio A.", 100, 18, 19),
    Task("task", "Optimización Consumo Energético (modo sleep)", "Antonio A.", 30, 20, 23,
         "Modem_PowerOff/Modem_Sleep implementados"),

    # ═══════════════════════════════════════════════════════════════════
    Task("phase", "📋 P1 — SISTEMA ENERGÉTICO (Prioridad Alta · ~2 meses)"),
    Task("task", "Validación Banco de Baterías Li-Ion", "Antonio A.", 0, 18, 19),
    Task("task", "Monitoreo de Estado de Baterías (ADC + sensores)", "Antonio A.", 0, 18, 20,
         "Voltaje, corriente, temperatura"),
    Task("task", "Monitoreo Sistema Fotovoltaico (panel solar)", "Antonio A.", 0, 19, 21),
    Task("task", "PCB Compacto Alimentación (dedicado)", "Antonio A.", 0, 18, 21),
    Task("task", "Protecciones Eléctricas (overvoltage, reverse, ESD)",
         "Antonio A.", 0, 19, 21),
    Task("task", "Validación Energética en Condiciones Reales", "Antonio A.", 0, 21, 23),

    # ═══════════════════════════════════════════════════════════════════
    Task("sep"),
    Task("phase", "📋 P2 — SEÑAL, DIAGNÓSTICO Y TELEMETRÍA (Prioridad Media)"),
    Task("task", "Métricas de Calidad de Señal (RSSI, SNR, RSRP)", "Antonio A.", 0, 20, 23),
    Task("task", "Telemetría Energética (envío datos batería/solar)", "Antonio A.", 0, 20, 24),
    Task("task", "Gestión Remota de Configuración", "Antonio A.", 15, 20, 24,
         "Modem_DownloadConfig + Apply_Remote_Config (parcial)"),
    Task("task", "Sistema de Alarmas Operativas", "Antonio A.", 0, 20, 23),
    Task("task", "Monitoreo Estado Interno (temperatura, watchdog, uptime)",
         "Antonio A.", 0, 20, 22),
    Task("task", "Compensación Térmica del Acelerómetro", "Antonio A.", 0, 20, 23),
    Task("task", "Caracterización de Enlace Celular", "Antonio A.", 0, 21, 24),
    Task("task", "Sistema Modular de Baterías (hot-swap)", "Antonio A.", 0, 22, 25),
    Task("task", "Gestión Energética Adaptativa (dynamic power mgmt)",
         "Antonio A.", 0, 22, 26),
    Task("task", "Optimización Bypass Energético", "Antonio A.", 0, 22, 26),
    Task("task", "Conectores Rápidos para Panel Solar", "Antonio A.", 0, 23, 26),

    # ═══════════════════════════════════════════════════════════════════
    Task("sep"),
    Task("phase", "📋 P3 — EVOLUCIÓN TECNOLÓGICA (Prioridad Baja)"),
    Task("task", "Arquitectura Multisensor I2C (expansión)", "Antonio A.", 0, 26, 29),
    Task("task", "Reinicio Remoto del Sistema", "Antonio A.", 0, 26, 29),
    Task("task", "Migración a Protocolo FTP para transmisión", "Antonio A.", 0, 26, 29),
    Task("task", "Interfaz USB-C para Configuración Local", "Antonio A.", 0, 27, 30),
    Task("task", "Salidas de Energía Modulares", "Antonio A.", 0, 27, 30),

    # ═══════════════════════════════════════════════════════════════════
    Task("sep"),
    Task("phase", "🧪 INTEGRACIÓN Y VALIDACIÓN"),
    Task("task", "Encapsulado Robusto (IP65 / cierre hermético)", "Antonio A.", 0, 19, 24),
    Task("task", "Validación contra Sistema de Referencia", "Antonio A.", 0, 20, 25,
         "Acelerómetro patrón"),
    Task("task", "Línea Base de Comportamiento Estructural", "Antonio A.", 0, 21, 25),
    Task("task", "Pruebas Operacionales en Terreno", "Antonio A.", 0, 23, 26),

    # ═══════════════════════════════════════════════════════════════════
    Task("sep"),
    Task("phase", "✅ LOGRADO — FASE 4: UPLOAD PIPELINE & ESTABILIDAD (Jun 2026)"),
    Task("task", "Sistema de Cola RAM Upload (8 slots circulares)",
         "Antonio A.", 100, 17, 18, "file_task + upload_queue_push/pop/peek"),
    Task("task", "Tracking Upload por Backup Registers RTC (BKP0R-BKP19R)",
         "Antonio A.", 100, 17, 18, "Survive NRST/IWDG, 608 índices max"),
    Task("task", "Modem_PowerOff reubicado a modem_task (NRST fix)",
         "Antonio A.", 100, 17, 18, "Evita reset post-upload, queue pop antes de poweroff"),
    Task("task", "Upload concurrente con adquisición (sensor no bloquea)",
         "Antonio A.", 100, 17, 18, "sensor_task adquiere mientras modem sube"),
    Task("task", "Boot scan con skip por backup register + .DONE fallback",
         "Antonio A.", 100, 17, 18, "file_task.c verifica BKP0R + .DONE file"),
    Task("task", "Fix upload_queue_pop(NULL,0) UB (memory corruption)",
         "Antonio A.", 100, 17, 18, "NULL guard en strncpy de file_task.c"),

    # ═══════════════════════════════════════════════════════════════════
    Task("sep"),
    Task("phase", "🏁 HITOS"),
    Task("milestone", "Sprint 01", "✓", 100, 0, 0),
    Task("milestone", "Sprint 02", "✓", 100, 1, 1),
    Task("milestone", "Análisis de Competencia y Costos", "✓", 100, 0, 0),
    Task("milestone", "FreeRTOS Migration Complete", "✓", 100, 16, 16),
    Task("milestone", "Test Suite Finalizada (47 tests)", "✓", 100, 17, 17),
    Task("milestone", "Upload Pipeline Estable (sin NRST)", "✓", 100, 17, 18,
         "Jun 2026 — 3 uploads encadenados exitosos"),
    Task("milestone", "P1 — Adquisición Continua", "Target", 0, 23, 23),
    Task("milestone", "P1 — Sistema Energético", "Target", 0, 23, 23),
    Task("milestone", "Validación Terreno", "Target", 0, 26, 26),
    Task("milestone", "Cierre Proyecto", "Target", 0, 30, 30,
         "Dic 2027 (P3 completo)"),
]


# ── Build the workbook ──────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Project schedule"

# Column widths
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 52
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 2
ws.column_dimensions["H"].width = 1

for mi in range(NUM_MONTHS):
    col = get_column_letter(9 + mi)  # I=9 → ...
    ws.column_dimensions[col].width = 5

# ── Title area (rows 1-3) ───────────────────────────────────────────────
ws.merge_cells("A1:H1")
ws["A1"].value = "CARTA GANTT 2026"
ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:H2")
ws["A2"].value = "HERMES A1 — Monitoreo Sísmico Estructural | FreeRTOS + CMSIS-RTOS v2 | STM32F446RE"
ws["A2"].font = subtitle_font

ws.merge_cells("A3:H3")
ws["A3"].value = "Antonio Avendaño Sanhueza | LIND SpA | Concepción, Chile | Junio 2026"
ws["A3"].font = subtitle_font

# ── Legend row (row 4) ──────────────────────────────────────────────────
row = 4
ws.row_dimensions[row].height = 20
ws.cell(row=row, column=2, value="Leyenda:").font = Font(bold=True, size=8)
legend_items = [
    (3, "✅ Completado", GREEN),
    (4, "En Progreso", BLUE),
    (5, "Iniciado", ORANGE),
    (6, "Planificado", LIGHT_GRAY),
    (7, "Hito", MILESTONE),
]
for col_idx, label, color in legend_items:
    c = ws.cell(row=row, column=col_idx)
    c.value = label
    c.font = Font(size=8)
    c.fill = PatternFill("solid", fgColor=color)
    c.border = thin_border

# ── Month header (row 5) ───────────────────────────────────────────────
row = 5
ws.row_dimensions[row].height = 22
ws.cell(row=row, column=1).value = ""
for mi, ml in enumerate(month_labels):
    col = 9 + mi
    c = ws.cell(row=row, column=col)
    c.value = ml
    c.font = Font(name="Calibri", bold=True, size=8, color=WHITE)
    c.fill = fills["darkblue"]
    c.alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom")
    c.border = thin_border

# ── Blank separator row (row 6) ─────────────────────────────────────────
row = 6
ws.row_dimensions[row].height = 8

# ── Data rows (row 7+) ─────────────────────────────────────────────────
def bar_color(pct):
    if pct >= 100:
        return "darkgreen"
    elif pct >= 75:
        return "green"
    elif pct >= 25:
        return "blue"
    elif pct > 0:
        return "orange"
    return None

def bar_fill(pct):
    b = bar_color(pct)
    return fills.get(b) if b else fills["lightgray"]

data_start = 7
r = data_start

# Column headers
ws.cell(row=r, column=2, value="Tarea").font = Font(bold=True, size=9)
ws.cell(row=r, column=3, value="Asignado").font = Font(bold=True, size=9)
ws.cell(row=r, column=4, value="Avance").font = Font(bold=True, size=9)
ws.cell(row=r, column=5, value="Inicio").font = Font(bold=True, size=9)
ws.cell(row=r, column=6, value="Fin").font = Font(bold=True, size=9)
for mi in range(NUM_MONTHS):
    c = ws.cell(row=r, column=9+mi)
    c.fill = fills["lightgray"]
    c.border = thin_border
r += 1

for t in ROWS:
    ws.row_dimensions[r].height = 22 if t.kind == "task" else 20

    # ── Phase row ────────────────────────────────────────────────────
    if t.kind == "phase":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2)
        c.value = t.label
        c.font = phase_font
        c.fill = fills["phase"]
        c.alignment = Alignment(vertical="center")
        for cc in [3,4,5,6]:
            ws.cell(row=r, column=cc).fill = fills["phase"]
            ws.cell(row=r, column=cc).border = thin_border
        # Fill all month columns with phase header colour
        for mi in range(NUM_MONTHS):
            ws.cell(row=r, column=9+mi).fill = fills["phase"]
            ws.cell(row=r, column=9+mi).border = thin_border
        r += 1
        continue

    # ── Separator ────────────────────────────────────────────────────
    if t.kind == "sep":
        for mi in range(NUM_MONTHS):
            ws.cell(row=r, column=9+mi).border = thin_border
        r += 1
        continue

    # ── Milestone ────────────────────────────────────────────────────
    if t.kind == "milestone":
        ws.cell(row=r, column=2).value = t.label
        ws.cell(row=r, column=2).font = milestone_font
        ws.cell(row=r, column=3).value = t.assigned
        ws.cell(row=r, column=3).font = Font(size=8, bold=(t.pct==100))
        ws.cell(row=r, column=4).value = f"{t.pct}%"
        ws.cell(row=r, column=4).font = pct_font
        # Month labels for start/end
        if t.sm < NUM_MONTHS:
            ws.cell(row=r, column=5).value = month_labels[t.sm]
        if t.em < NUM_MONTHS:
            ws.cell(row=r, column=6).value = month_labels[t.em]
        ws.cell(row=r, column=5).font = date_font
        ws.cell(row=r, column=6).font = date_font
        # Highlight month
        for mi in range(NUM_MONTHS):
            c = ws.cell(row=r, column=9+mi)
            c.border = thin_border
            if t.sm <= mi <= t.em:
                c.fill = fills["milestone"]
                c.value = "◆" if t.pct >= 100 else "◇"
                c.font = Font(size=10, bold=True)
                c.alignment = Alignment(horizontal="center")
        r += 1
        continue

    # ── Task ─────────────────────────────────────────────────────────
    ws.cell(row=r, column=2).value = t.label
    ws.cell(row=r, column=2).font = task_font
    ws.cell(row=r, column=3).value = t.assigned
    ws.cell(row=r, column=3).font = Font(size=8)
    ws.cell(row=r, column=4).value = f"{t.pct}%"
    ws.cell(row=r, column=4).font = pct_font
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    if t.sm < NUM_MONTHS:
        ws.cell(row=r, column=5).value = month_labels[t.sm]
    if t.em < NUM_MONTHS:
        ws.cell(row=r, column=6).value = month_labels[t.em]
    ws.cell(row=r, column=5).font = date_font
    ws.cell(row=r, column=6).font = date_font

    # Bar colouring
    bf = bar_fill(t.pct)
    for mi in range(NUM_MONTHS):
        c = ws.cell(row=r, column=9+mi)
        c.border = thin_border
        if t.sm <= mi <= t.em:
            c.fill = bf
        else:
            c.fill = fills["white"]

    r += 1

# ── Final row: note ────────────────────────────────────────────────────
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=24)
ws.cell(row=r, column=2).value = (
    "Nota: Avance actualizado al 24 de Junio 2026. "
    "Fase 4 completada: upload pipeline estable, NRST eliminado, backup registers funcionando. "
    "3 uploads encadenados exitosos (TRIG_006→007→008). "
    "Colores: Verde = 100%  |  Azul = 50-99%  |  Naranjo = 1-49%  |  Sin relleno = 0%"
)
ws.cell(row=r, column=2).font = Font(size=8, italic=True, color=DARK_GRAY)

# ── Freeze panes ────────────────────────────────────────────────────────
ws.freeze_panes = "I8"

# ── Print setup ─────────────────────────────────────────────────────────
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.paperSize = ws.PAPERSIZE_A3

# ── Save ────────────────────────────────────────────────────────────────
wb.save(DST)
print(f"✅ Saved → {DST}")
print(f"   Sheet: {ws.title}")
print(f"   Timelines: {NUM_MONTHS} columns")
print(f"   Tasks/rows: {len(ROWS)}")
