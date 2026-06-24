#!/usr/bin/env python3
"""Fix week ordering in rebuilt planning xlsx — sort by actual date descending."""
import openpyxl, glob, os, re, sys
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

PLANNING_DIR = os.path.dirname(os.path.abspath(__file__))
matches = glob.glob(os.path.join(PLANNING_DIR, 'Planif*'))
src = matches[0]

wb = openpyxl.load_workbook(src)
ws = wb.active
print('Loaded: %d rows, %d cols' % (ws.max_row, ws.max_column))

# Extract week data: each week is 7 rows (header + subheader + 4 items + total)
weeks = []
r = 7  # data starts at row 7
while r <= ws.max_row:
    title = ws.cell(r, 1).value
    if title and 'Semana' in str(title):
        week = {'title': str(title), 'start_row': r, 'rows': []}
        for i in range(7):  # 7 rows per week block
            if r + i <= ws.max_row:
                row_data = []
                for c in range(1, 6):
                    row_data.append(ws.cell(r + i, c).value)
                week['rows'].append(row_data)
        weeks.append(week)
        r += 7
    elif title and ('TAREAS' in str(title) or 'PROGRESO' in str(title) or 'FUTURAS' in str(title)):
        break
    else:
        r += 1

print('Found %d week blocks' % len(weeks))

# Parse date from title to sort
def parse_week_date(title):
    """Extract start date from 'Semana NN: DD al DD Mes YYYY'"""
    # Handle "Semana 29-30: 14 al 25 Jul 2026"
    m = re.search(r'(\d{1,2})\s+al\s+(\d{1,2})\s+(\w+)\s+(\d{4})', title)
    if m:
        day_start = int(m.group(1))
        month_str = m.group(3)
        year = int(m.group(4))
        months_es = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
                     'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}
        month = months_es.get(month_str, 1)
        return datetime(year, month, day_start)
    # Handle "Semana 14: 01 al 02 Abril" (no year = 2026)
    m = re.search(r'(\d{1,2})\s+al\s+(\d{1,2})\s+(\w+)', title)
    if m:
        day_start = int(m.group(1))
        month_str = m.group(3)
        months_es = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
                     'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}
        month = months_es.get(month_str, 1)
        return datetime(2026, month, day_start)
    return datetime(2025, 1, 1)

# Identify current week
current = None
past_future = []
for w in weeks:
    if '[ACTUAL]' in w['title']:
        current = w
    else:
        past_future.append(w)

# Sort non-current by date DESCENDING (most recent first)
past_future.sort(key=lambda w: parse_week_date(w['title']), reverse=True)

# Final order: current first, then sorted past/future
ordered = []
if current:
    ordered.append(current)
ordered.extend(past_future)

print('\nFinal order:')
for i, w in enumerate(ordered):
    tag = ' [ACTUAL]' if '[ACTUAL]' in w['title'] else ''
    print('  %d. %s%s' % (i+1, w['title'][:50], tag))

# Now rebuild the worksheet with correct order
# Keep rows 1-6 as title + column headers
# Rewrite rows 7+ with ordered weeks
# First, collect everything after weeks (TAREAS etc.)
after_weeks = []
r = 7 + len(weeks) * 7
while r <= ws.max_row:
    row_data = []
    for c in range(1, 6):
        row_data.append(ws.cell(r, c).value)
    after_weeks.append(row_data)
    r += 1

# Clear all data rows
for r in range(7, ws.max_row + 1):
    for c in range(1, 6):
        ws.cell(r, c).value = None

# Write ordered weeks
r = 7
for week in ordered:
    for row_data in week['rows']:
        for c, val in enumerate(row_data):
            ws.cell(r, c + 1).value = val
        r += 1
    # Add spacing row
    r += 1

# Write remaining content (TAREAS etc.)
for row_data in after_weeks:
    for c, val in enumerate(row_data):
        ws.cell(r, c + 1).value = val
    r += 1

wb.save(src)
print('\nFixed! Saved %d rows' % r)
