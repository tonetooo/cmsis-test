#!/usr/bin/env python3
"""Fix row heights in planning xlsx so June+ rows are visible."""
import openpyxl, glob, os, sys
sys.stdout.reconfigure(encoding='utf-8')

PLANNING_DIR = os.path.dirname(os.path.abspath(__file__))
matches = glob.glob(os.path.join(PLANNING_DIR, 'Planif*'))
if not matches:
    raise RuntimeError("No planning file found")

src = matches[0]

# Read
wb = openpyxl.load_workbook(src)
ws = wb.active
print('Loaded: %d rows, %d cols' % (ws.max_row, ws.max_column))

# Identify section structure
# Each week block has: header row (13h), subheader row (15.75h or 13h), 
# 4 item rows (60h for May, 13h for June), total row (13h)

# Fix: For all rows after row 65, set proper heights
# May pattern: header=13, subheader=15.75, items=60, total=13
# June pattern: header=13, subheader=13, items=13, total=13 (ALL TINY)

changes = 0
for r in range(65, ws.max_row + 1):
    cell_b = ws.cell(r, 2).value
    cell_b_str = str(cell_b) if cell_b else ''
    
    row_dim = ws.row_dimensions.get(r)
    current_h = row_dim.height if row_dim else None
    
    # Detect row type by content
    if 'Semana' in cell_b_str or 'RESUMEN' in cell_b_str:
        # Section header - keep at 13-15
        target = 15
    elif cell_b == 'Total HH:':
        target = 13
    elif cell_b_str in ('1', '2', '3', '4', '5') and r > 59:
        # Item row - these should be taller for readability
        # Check if the row has actual content in columns C-H
        has_content = False
        for c in range(3, 9):
            v = ws.cell(r, c).value
            if v and str(v).strip() and str(v).strip() != '---':
                has_content = True
                break
        if has_content:
            target = 45  # Compact but readable
        else:
            target = 25  # Less content = smaller
    elif cell_b_str == 'TAREAS EN PROGRESO' or cell_b_str == 'TAREAS FUTURAS':
        target = 20
    elif cell_b == 'Ítem':
        # Subheader
        target = 15
    else:
        target = None  # Don't change
    
    if target is not None and current_h != target:
        ws.row_dimensions[r].height = target
        changes += 1
        print('Row %d (%s): height %.1f -> %d' % (r, cell_b_str[:25], current_h or 0, target))

# Save
wb.save(src)
print('\nDone: %d row heights fixed' % changes)
print('File: %s' % os.path.basename(src))
