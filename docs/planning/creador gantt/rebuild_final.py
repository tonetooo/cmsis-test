#!/usr/bin/env python3
"""
HERMES-A1 Planificacion Semanal — FINAL REBUILD
Reads old data, creates fresh xlsx with:
- Reverse chronological order (most recent week first, ACTUAL marked)
- Max 15 HH/week enforced
- Beautiful formatting, proper column widths, visible text
- No merged cells in data area (prevents row reorder issues)
"""
import openpyxl, glob, os, re, sys
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

PLANNING_DIR = os.path.dirname(os.path.abspath(__file__))
matches = glob.glob(os.path.join(PLANNING_DIR, 'Planif*'))
if not matches:
    raise RuntimeError("No planning file found")

# ═══════════════════════════════════════════════════════════════════════
# STEP 1: Read source data from old file
# ═══════════════════════════════════════════════════════════════════════
# Source: the restored original (has all 17 weeks with full data)
src = os.path.join(PLANNING_DIR, 'Planificacion_ORIGINAL.xlsx')
DST_USER = os.path.join(PLANNING_DIR, 'Planificaci\u00f3n semanal actual .xlsx')

print('Source: %s (%d bytes)' % (os.path.basename(src), os.path.getsize(src)))
wb_old = openpyxl.load_workbook(src)
ws_old = wb_old.active

months_es = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'Abril':4,'May':5,'Jun':6,
             'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}

def parse_week_date(title):
    months_re = '|'.join(months_es.keys())
    # Pattern 1: "DD Mon al" — first date in cross-month range like "30 Jun al 04 Jul 2026"
    m = re.search(r'(\d{1,2})\s+(' + months_re + r')\s+al\s', title)
    if m:
        day, month_s = int(m.group(1)), m.group(2)
        ym = re.search(r'(\d{4})', title)
        year = int(ym.group(1)) if ym else 2026
        return datetime(year, months_es.get(month_s, 1), day)
    # Pattern 2: "DD al DD Mon [YYYY]" — same month like "01 al 02 Abril" or "09 al 13 Jun 2026"
    m = re.search(r'(\d{1,2})\s+al\s+(\d{1,2})\s+(\w+)', title)
    if m:
        day, month_s = int(m.group(1)), m.group(3)
        ym = re.search(r'(\d{4})', title)
        year = int(ym.group(1)) if ym else 2026
        return datetime(year, months_es.get(month_s, 1), day)
    return datetime(2025, 1, 1)

def parse_hh(text):
    m = re.search(r'\[(\d+)\s*HH\]', str(text or ''))
    return int(m.group(1)) if m else 0

# Extract week blocks
weeks = []
current_week = None
section_break = False

for r in range(2, ws_old.max_row + 1):
    b_val = ws_old.cell(r, 2).value
    b_str = str(b_val) if b_val else ''
    
    if 'Semana' in b_str and ':' in b_str:
        if current_week:
            weeks.append(current_week)
        current_week = {
            'title': b_str.strip(),
            'items': [],
        }
    elif 'TAREAS' in b_str or 'RESUMEN' in b_str or 'Planificacion' in b_str or 'Proximas' in b_str:
        # Section break — close current week but DON'T stop scanning
        if current_week:
            weeks.append(current_week)
            current_week = None
        continue  # skip this row, keep scanning for more Semana blocks
    elif current_week is not None:
        if b_str in ('1', '2', '3', '4', '5'):
            item = {
                'num': b_str,
                'pasada': str(ws_old.cell(r, 3).value or '').strip(),
                'pendiente': str(ws_old.cell(r, 5).value or '').strip(),
                'motivo': str(ws_old.cell(r, 6).value or '').strip(),
                'esta_semana': str(ws_old.cell(r, 8).value or '').strip(),
            }
            if item['pasada'] in ('', 'None'): item['pasada'] = '---'
            if item['pendiente'] in ('', 'None'): item['pendiente'] = '---'
            if item['motivo'] in ('', 'None'): item['motivo'] = '---'
            if item['esta_semana'] in ('', 'None'): item['esta_semana'] = '---'
            current_week['items'].append(item)

if current_week:
    weeks.append(current_week)

# Extract TAREAS
tareas_progreso = []
tareas_futuras = []
in_prog = False
in_fut = False
for r in range(65, 76):
    b_val = ws_old.cell(r, 2).value
    c_val = str(ws_old.cell(r, 3).value or '').strip()
    b_str = str(b_val) if b_val else ''
    if b_str == 'TAREAS EN PROGRESO':
        in_prog = True; in_fut = False; continue
    elif b_str == 'TAREAS FUTURAS':
        in_fut = True; in_prog = False; continue
    elif b_str in ('1','2','3','4','5'):
        if in_prog and len(tareas_progreso) < 5:
            tareas_progreso.append(c_val)
        elif in_fut and len(tareas_futuras) < 5:
            tareas_futuras.append(c_val)

print('Extracted %d weeks, %d tareas prog, %d tareas fut' % (len(weeks), len(tareas_progreso), len(tareas_futuras)))

# ═══════════════════════════════════════════════════════════════════════
# STEP 2: Fix hours (max 15 HH/week)
# ═══════════════════════════════════════════════════════════════════════
def fix_hours(week, max_hh=15):
    total = sum(parse_hh(item['esta_semana']) for item in week['items'])
    if total <= max_hh or total == 0:
        return
    scale = max_hh / total
    new_total = 0
    for item in week['items']:
        h = parse_hh(item['esta_semana'])
        if h > 0:
            new_h = max(1, round(h * scale))
            item['esta_semana'] = re.sub(r'\[\d+\s*HH\]', '[%d HH]' % new_h, item['esta_semana'])
            new_total += new_h
    # Adjust last non-zero item for exact total
    diff = max_hh - new_total
    if diff != 0:
        for item in reversed(week['items']):
            h = parse_hh(item['esta_semana'])
            if h + diff > 0:
                new_h = h + diff
                item['esta_semana'] = re.sub(r'\[\d+\s*HH\]', '[%d HH]' % new_h, item['esta_semana'])
                break

for w in weeks:
    fix_hours(w)

# ═══════════════════════════════════════════════════════════════════════
# STEP 3: Sort weeks (most recent first, [ACTUAL] at top)
# ═══════════════════════════════════════════════════════════════════════
current = None
others = []
for w in weeks:
    if '[ACTUAL]' in w['title']:
        current = w
    else:
        # Auto-mark Semana 26 as current (week 24-27 Jun 2026 = ACTUAL)
        if '23 al 27 Jun' in w['title'] or '24 al 27 Jun' in w['title']:
            w['title'] = w['title'].replace('2026', '2026 [ACTUAL]') if '2026' in w['title'] else w['title'] + ' [ACTUAL]'
            current = w
        else:
            others.append(w)

others.sort(key=lambda w: parse_week_date(w['title']), reverse=True)

ordered = []
if current:
    ordered.append(current)
ordered.extend(others)

# ═══════════════════════════════════════════════════════════════════════
# STEP 4: Create fresh workbook with beautiful formatting
# ═══════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Planificacion Semanal"

# Column widths — all text visible
ws.column_dimensions['A'].width = 4    # Item number
ws.column_dimensions['B'].width = 48   # Semana pasada
ws.column_dimensions['C'].width = 38   # Pendientes
ws.column_dimensions['D'].width = 28   # Motivo
ws.column_dimensions['E'].width = 55   # Esta semana

# Colors
DK_BLUE = "1F3864"
MD_BLUE = "2E75B6"
LT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LT_GRAY = "F2F2F2"
DK_GRAY = "404040"
GREEN = "70AD47"
LT_GREEN = "E2EFDA"
ORANGE = "ED7D31"
LT_ORANGE = "FCE4D6"
AMBER = "FFC000"
LT_AMBER = "FFF2CC"
RED = "C00000"
ACCENT = "5B9BD5"

# Fills
f_title = PatternFill("solid", fgColor=DK_BLUE)
f_subtitle = PatternFill("solid", fgColor="2B4C7E")
f_colhdr = PatternFill("solid", fgColor=MD_BLUE)
f_week = PatternFill("solid", fgColor=DK_BLUE)
f_week_actual = PatternFill("solid", fgColor=GREEN)
f_item1 = PatternFill("solid", fgColor=WHITE)
f_item2 = PatternFill("solid", fgColor=LT_BLUE)
f_total = PatternFill("solid", fgColor=LT_GRAY)
f_sep = PatternFill("solid", fgColor=LT_GRAY)
f_tprog_hdr = PatternFill("solid", fgColor=ORANGE)
f_tprog = PatternFill("solid", fgColor=LT_ORANGE)
f_tfut_hdr = PatternFill("solid", fgColor=AMBER)
f_tfut = PatternFill("solid", fgColor=LT_AMBER)

# Fonts
fn_title = Font(name="Calibri", bold=True, size=16, color=WHITE)
fn_subtitle = Font(name="Calibri", size=11, color=WHITE)
fn_note = Font(name="Calibri", italic=True, size=9, color=ACCENT)
fn_week = Font(name="Calibri", bold=True, size=12, color=WHITE)
fn_week_act = Font(name="Calibri", bold=True, size=12, color=WHITE)
fn_colhdr = Font(name="Calibri", bold=True, size=10, color=WHITE)
fn_item_num = Font(name="Calibri", bold=True, size=10, color=DK_BLUE)
fn_item_txt = Font(name="Calibri", size=10, color=DK_GRAY)
fn_item_bold = Font(name="Calibri", size=10, color=DK_GRAY, bold=True)
fn_total = Font(name="Calibri", bold=True, size=10, color=DK_GRAY)
fn_total_over = Font(name="Calibri", bold=True, size=10, color=RED)
fn_tarea_hdr = Font(name="Calibri", bold=True, size=12, color=WHITE)
fn_tarea_txt = Font(name="Calibri", size=10, color=DK_GRAY)
fn_footer = Font(name="Calibri", italic=True, size=9, color=DK_GRAY)

# Borders
thin = Side(style="thin", color="B4B4B4")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
bd_bottom = Border(bottom=Side(style="medium", color=DK_BLUE))

# Alignment
ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
al = Alignment(horizontal="left", vertical="center", wrap_text=True)
alt = Alignment(horizontal="left", vertical="top", wrap_text=True)

def fill_row(ws, row, fill, cols=5):
    for c in range(1, cols+1):
        ws.cell(row, c).fill = fill
        ws.cell(row, c).border = bd

# ── Row 1: Title ──────────────────────────────────────────────────
r = 1
ws.merge_cells('A1:E1')
c = ws.cell(r, 1, value="HERMES-A1 — Planificacion Semanal")
c.font = fn_title; c.fill = f_title; c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 38
fill_row(ws, r, f_title)

# ── Row 2: Subtitle ──────────────────────────────────────────────
r = 2
ws.merge_cells('A2:E2')
c = ws.cell(r, 1, value="Monitoreo Sismico Estructural | FreeRTOS + STM32F446RE | LIND SpA | Junio 2026")
c.font = fn_subtitle; c.fill = f_title; c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 22
fill_row(ws, r, f_title)

# ── Row 3: Max HH note ───────────────────────────────────────────
r = 3
ws.merge_cells('A3:E3')
c = ws.cell(r, 1, value="Maximo: 15 HH/semana | Orden cronologico inverso (mas reciente primero) | Sesion actual: [ACTUAL]")
c.font = fn_note; c.fill = f_subtitle; c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 18
fill_row(ws, r, f_subtitle)

# ── Row 4: blank ──────────────────────────────────────────────────
r = 4
ws.row_dimensions[r].height = 6
fill_row(ws, r, f_sep)

# ── Row 5: Column headers ────────────────────────────────────────
r = 5
ws.row_dimensions[r].height = 24
hdrs = [(1, "#"), (2, "Actividades Semana Pasada"), (3, "Actividades Pendientes"), (4, "Motivo / Dependencia"), (5, "Actividades para Esta Semana")]
for col, txt in hdrs:
    c = ws.cell(r, col, value=txt)
    c.font = fn_colhdr; c.fill = f_colhdr; c.alignment = ac; c.border = bd

# ── Row 6: blank ──────────────────────────────────────────────────
r = 6
ws.row_dimensions[r].height = 4

# ═══════════════════════════════════════════════════════════════════════
# STEP 5: Write week blocks
# ═══════════════════════════════════════════════════════════════════════
r = 7

for wi, week in enumerate(ordered):
    is_actual = '[ACTUAL]' in week['title']
    title_clean = week['title'].replace('[ACTUAL]', '').strip()
    if is_actual:
        title_clean += '  [SESION ACTUAL]'
    
    # Week header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    hdr_fill = f_week_actual if is_actual else f_week
    c = ws.cell(r, 1, value=title_clean)
    c.font = fn_week_act if is_actual else fn_week
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    fill_row(ws, r, hdr_fill)
    ws.row_dimensions[r].height = 28
    r += 1
    
    # Items
    total_hh = sum(parse_hh(item['esta_semana']) for item in week['items'])
    for ii, item in enumerate(week['items']):
        ifill = f_item1 if ii % 2 == 0 else f_item2
        
        # #
        c = ws.cell(r, 1, value=item['num'])
        c.font = fn_item_num; c.fill = ifill; c.alignment = ac; c.border = bd
        
        # Semana pasada
        c = ws.cell(r, 2, value=item['pasada'])
        c.font = fn_item_txt; c.fill = ifill; c.alignment = alt; c.border = bd
        
        # Pendientes
        c = ws.cell(r, 3, value=item['pendiente'])
        c.font = fn_item_txt; c.fill = ifill; c.alignment = alt; c.border = bd
        
        # Motivo
        c = ws.cell(r, 4, value=item['motivo'])
        c.font = fn_item_txt; c.fill = ifill; c.alignment = alt; c.border = bd
        
        # Esta semana
        est = item['esta_semana']
        c = ws.cell(r, 5, value=est)
        c.font = fn_item_bold if parse_hh(est) > 0 else fn_item_txt
        c.fill = ifill; c.alignment = alt; c.border = bd
        
        ws.row_dimensions[r].height = 52
        r += 1
    
    # Total row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    hh_fn = fn_total if total_hh <= 15 else fn_total_over
    c = ws.cell(r, 1, value="Total HH: %d / 15" % total_hh)
    c.font = hh_fn; c.fill = f_total
    c.alignment = Alignment(horizontal="right", vertical="center")
    for cc in range(1, 6):
        ws.cell(r, cc).fill = f_total; ws.cell(r, cc).border = bd
    ws.row_dimensions[r].height = 20
    r += 1
    
    # Separator between weeks
    if wi < len(ordered) - 1:
        ws.row_dimensions[r].height = 6
        fill_row(ws, r, f_sep)
        r += 1

# ═══════════════════════════════════════════════════════════════════════
# STEP 6: TAREAS EN PROGRESO
# ═══════════════════════════════════════════════════════════════════════
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, value="TAREAS EN PROGRESO")
c.font = fn_tarea_hdr; c.fill = f_tprog_hdr; c.alignment = Alignment(horizontal="left", vertical="center")
fill_row(ws, r, f_tprog_hdr)
ws.row_dimensions[r].height = 28
r += 1

for i, task in enumerate(tareas_progreso):
    c = ws.cell(r, 1, value=str(i+1))
    c.font = fn_item_num; c.fill = f_tprog; c.alignment = ac; c.border = bd
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    c = ws.cell(r, 2, value=task)
    c.font = fn_tarea_txt; c.fill = f_tprog; c.alignment = al; c.border = bd
    for cc in range(2, 6):
        ws.cell(r, cc).fill = f_tprog; ws.cell(r, cc).border = bd
    ws.row_dimensions[r].height = 32
    r += 1

# ═══════════════════════════════════════════════════════════════════════
# STEP 7: TAREAS FUTURAS
# ═══════════════════════════════════════════════════════════════════════
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, value="TAREAS FUTURAS")
c.font = fn_tarea_hdr; c.fill = f_tfut_hdr; c.alignment = Alignment(horizontal="left", vertical="center")
fill_row(ws, r, f_tfut_hdr)
ws.row_dimensions[r].height = 28
r += 1

for i, task in enumerate(tareas_futuras):
    c = ws.cell(r, 1, value=str(i+1))
    c.font = fn_item_num; c.fill = f_tfut; c.alignment = ac; c.border = bd
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    c = ws.cell(r, 2, value=task)
    c.font = fn_tarea_txt; c.fill = f_tfut; c.alignment = al; c.border = bd
    for cc in range(2, 6):
        ws.cell(r, cc).fill = f_tfut; ws.cell(r, cc).border = bd
    ws.row_dimensions[r].height = 32
    r += 1

# ═══════════════════════════════════════════════════════════════════════
# STEP 8: Footer
# ═══════════════════════════════════════════════════════════════════════
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(r, 1, value="Generado automaticamente desde GANTT 2026 | Actualizado: 24 Junio 2026 | LIND SpA")
c.font = fn_footer; c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 18

# ═══════════════════════════════════════════════════════════════════════
# STEP 9: Freeze + print setup
# ═══════════════════════════════════════════════════════════════════════
ws.freeze_panes = "A6"
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.paperSize = ws.PAPERSIZE_A3

# ═══════════════════════════════════════════════════════════════════════
# STEP 10: Save
# ═══════════════════════════════════════════════════════════════════════
import shutil

# Save to temp first
tmp = DST_USER + '.tmp.xlsx'
wb.save(tmp)
print('Saved temp: %d bytes' % os.path.getsize(tmp))

# Replace user file
try: os.remove(DST_USER)
except (PermissionError, FileNotFoundError): pass

shutil.move(tmp, DST_USER)
print('Final: %s (%d bytes)' % (os.path.basename(DST_USER), os.path.getsize(DST_USER)))
print('Rows: %d, Weeks: %d' % (r, len(ordered)))

print('\nOrder (most recent first):')
for i, w in enumerate(ordered):
    total = sum(parse_hh(item['esta_semana']) for item in w['items'])
    tag = ' [ACTUAL]' if '[ACTUAL]' in w['title'] else ''
    status = 'OK' if total <= 15 else 'OVER %d' % total
    print('  %2d. %s: %d HH [%s]%s' % (i+1, w['title'][:45], total, status, tag))
