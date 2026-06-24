#!/usr/bin/env python3
import openpyxl, glob, sys
sys.stdout.reconfigure(encoding='utf-8')
matches = glob.glob('Planif*')
wb = openpyxl.load_workbook(matches[0])
ws = wb.active
# Print all columns for rows 83-110 (Semana 23-26) to see actual content
for r in range(83, 111):
    vals = []
    for c in range(1, 10):
        v = ws.cell(r, c).value
        if v is not None:
            vals.append('C%d=%s' % (c, str(v)[:50]))
    if vals:
        print('Row %d: %s' % (r, '  |  '.join(vals)))
# Also check rows 65-75 (TAREAS EN PROGRESO and TAREAS FUTURAS)
print("\n--- TAREAS EN PROGRESO (rows 65-68) ---")
for r in range(65, 69):
    vals = []
    for c in range(1, 10):
        v = ws.cell(r, c).value
        if v is not None:
            vals.append('C%d=%s' % (c, str(v)[:80]))
    print('Row %d: %s' % (r, '  |  '.join(vals) if vals else '(empty)'))

print("\n--- TAREAS FUTURAS (rows 70-75) ---")
for r in range(70, 76):
    vals = []
    for c in range(1, 10):
        v = ws.cell(r, c).value
        if v is not None:
            vals.append('C%d=%s' % (c, str(v)[:80]))
    print('Row %d: %s' % (r, '  |  '.join(vals) if vals else '(empty)'))
